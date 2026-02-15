"""
电动汽车知识问答系统 - 文档处理服务
"""
import os
import tempfile
from pathlib import Path
from typing import List, Dict, Any, Optional
from loguru import logger

import PyPDF2
from docx import Document
import openpyxl
from unstructured.partition.auto import partition

from app.core.config import settings
from app.services.rag_service import rag_service

class EVDocumentService:
    """电动汽车领域文档处理服务"""
    
    def __init__(self):
        self.allowed_extensions = settings.ALLOWED_EXTENSIONS
        self.max_file_size = settings.MAX_FILE_SIZE
        
    def process_document(self, file_path: str, metadata: Dict = None) -> Dict[str, Any]:
        """处理单个文档"""
        try:
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext not in self.allowed_extensions:
                return {
                    "success": False,
                    "error": f"不支持的文件格式: {file_ext}",
                    "allowed_formats": self.allowed_extensions
                }
            
            # 检查文件大小
            file_size = os.path.getsize(file_path)
            if file_size > self.max_file_size:
                return {
                    "success": False,
                    "error": f"文件过大: {file_size}字节 (最大: {self.max_file_size}字节)"
                }
            
            # 提取文本内容
            text_content = self._extract_text(file_path, file_ext)
            
            if not text_content or len(text_content.strip()) < 10:
                return {
                    "success": False,
                    "error": "文档内容为空或过短"
                }
            
            # 增强电动汽车领域内容
            enhanced_content = self._enhance_ev_content(text_content)
            
            # 准备元数据
            doc_metadata = {
                "filename": Path(file_path).name,
                "extension": file_ext,
                "size": file_size,
                "domain": "electric_vehicles",
                "processed": True
            }
            
            if metadata:
                doc_metadata.update(metadata)
            
            # 添加到RAG知识库
            success = rag_service.add_documents([enhanced_content], [doc_metadata])
            
            if success:
                return {
                    "success": True,
                    "filename": Path(file_path).name,
                    "content_length": len(text_content),
                    "enhanced_length": len(enhanced_content),
                    "metadata": doc_metadata,
                    "domain": "electric_vehicles"
                }
            else:
                return {
                    "success": False,
                    "error": "添加到知识库失败"
                }
                
        except Exception as e:
            logger.error(f"❌ 文档处理失败: {e}")
            return {
                "success": False,
                "error": f"处理失败: {str(e)}"
            }
    
    def _extract_text(self, file_path: str, file_ext: str) -> str:
        """根据文件类型提取文本"""
        try:
            if file_ext == '.pdf':
                return self._extract_pdf(file_path)
            elif file_ext == '.docx':
                return self._extract_docx(file_path)
            elif file_ext == '.txt':
                return self._extract_txt(file_path)
            elif file_ext in ['.xlsx', '.xls']:
                return self._extract_excel(file_path)
            elif file_ext == '.md':
                return self._extract_markdown(file_path)
            else:
                # 使用unstructured作为后备方案
                return self._extract_with_unstructured(file_path)
                
        except Exception as e:
            logger.error(f"❌ 文本提取失败 ({file_ext}): {e}")
            return ""
    
    def _extract_pdf(self, file_path: str) -> str:
        """提取PDF文本"""
        text = ""
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page_num in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[page_num]
                text += page.extract_text() + "\n\n"
        return text
    
    def _extract_docx(self, file_path: str) -> str:
        """提取DOCX文本"""
        doc = Document(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text
    
    def _extract_txt(self, file_path: str) -> str:
        """提取TXT文本"""
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.read()
    
    def _extract_excel(self, file_path: str) -> str:
        """提取Excel文本"""
        text = ""
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"工作表: {sheet_name}\n"
            
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) for cell in row if cell)
                if row_text:
                    text += row_text + "\n"
            
            text += "\n"
        
        return text
    
    def _extract_markdown(self, file_path: str) -> str:
        """提取Markdown文本"""
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.read()
    
    def _extract_with_unstructured(self, file_path: str) -> str:
        """使用unstructured提取文本"""
        elements = partition(filename=file_path)
        text = "\n".join(str(element) for element in elements)
        return text
    
    def _enhance_ev_content(self, content: str) -> str:
        """增强电动汽车领域内容"""
        enhanced = content
        
        # 添加领域标识
        enhanced = f"[电动汽车领域文档]\n{enhanced}"
        
        # 检查是否包含电动汽车关键词
        ev_keywords = settings.DOMAIN_KEYWORDS
        found_keywords = [kw for kw in ev_keywords if kw in content]
        
        if found_keywords:
            enhanced += f"\n\n[关键词]: {', '.join(found_keywords)}"
        
        return enhanced
    
    def batch_process(self, file_paths: List[str], metadata_list: List[Dict] = None) -> Dict[str, Any]:
        """批量处理文档"""
        results = {
            "total": len(file_paths),
            "successful": 0,
            "failed": 0,
            "details": []
        }
        
        for i, file_path in enumerate(file_paths):
            metadata = metadata_list[i] if metadata_list and i < len(metadata_list) else None
            
            result = self.process_document(file_path, metadata)
            result["file_index"] = i
            result["file_path"] = file_path
            
            results["details"].append(result)
            
            if result.get("success"):
                results["successful"] += 1
            else:
                results["failed"] += 1
        
        return results
    
    def validate_document(self, file_path: str) -> Dict[str, Any]:
        """验证文档"""
        try:
            file_ext = Path(file_path).suffix.lower()
            file_size = os.path.getsize(file_path)
            
            is_valid = (
                file_ext in self.allowed_extensions and
                file_size <= self.max_file_size and
                os.path.exists(file_path)
            )
            
            # 预览前100个字符
            preview = ""
            if is_valid and file_ext == '.txt':
                with open(file_path, 'r', encoding='utf-8') as f:
                    preview = f.read(100)
            elif is_valid:
                # 尝试提取预览
                try:
                    text = self._extract_text(file_path, file_ext)
                    preview = text[:100]
                except:
                    preview = "[无法预览]"
            
            return {
                "valid": is_valid,
                "filename": Path(file_path).name,
                "extension": file_ext,
                "size": file_size,
                "max_size": self.max_file_size,
                "allowed": file_ext in self.allowed_extensions,
                "preview": preview,
                "domain_suitable": self._check_domain_suitability(file_path)
            }
            
        except Exception as e:
            logger.error(f"❌ 文档验证失败: {e}")
            return {
                "valid": False,
                "error": str(e)
            }
    
    def _check_domain_suitability(self, file_path: str) -> Dict[str, Any]:
        """检查文档是否适合电动汽车领域"""
        try:
            # 提取部分文本进行检查
            file_ext = Path(file_path).suffix.lower()
            sample_text = self._extract_text(file_path, file_ext)[:1000]
            
            ev_keywords = settings.DOMAIN_KEYWORDS
            found_keywords = []
            
            for keyword in ev_keywords:
                if keyword in sample_text:
                    found_keywords.append(keyword)
            
            relevance_score = len(found_keywords) / len(ev_keywords) * 100
            
            return {
                "relevant": len(found_keywords) > 0,
                "found_keywords": found_keywords,
                "relevance_score": round(relevance_score, 1),
                "sample_text": sample_text[:200] + "..." if sample_text else ""
            }
            
        except Exception as e:
            logger.error(f"❌ 领域适合性检查失败: {e}")
            return {
                "relevant": False,
                "error": str(e)
            }

# 全局文档服务实例
document_service = EVDocumentService()